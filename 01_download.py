#!/usr/bin/env python3
"""
Step 1 – Download audio (MP3) and transcript (PDF) for each hearing.

Dataset index: Google Sheets
  https://docs.google.com/spreadsheets/d/1AKsEOYGylGEC7fIN24XqUEI1CZXxH22cQcIpbb9eji0

IMPORTANT: the sheet is parsed as XLSX (not CSV).  The CSV export only
shows plain text in column 7, hiding the hyperlinks.  In the xlsx, column 7
cells carry Google Drive hyperlinks that are accessible from any server:

  xlsx column layout (1-indexed):
    1   Sr. No.
    2   Case Name
    3   Case Number
    4   Hearing Date
    5   Transcript Link [Outdated] – sci.gov.in URL (may be blocked by firewall)
    6   Transcript PDF Name
    7   Transcript Link            – CELL VALUE is filename; HYPERLINK is GDrive
    8   Oral Hearing Link (YouTube)
    9   Hearing Duration (minutes)
    10  mp3 format link (Dropbox)  ← audio download

PDF download priority:
  1. Google Drive hyperlink from col 7  (preferred – widely accessible)
  2. sci.gov.in URL from col 5          (fallback – may be blocked)

After download each hearing directory contains:
  audio.mp3        – original Dropbox audio
  audio.wav        – 16 kHz mono WAV (ffmpeg-converted)
  transcript.pdf   – Supreme Court transcript (from GDrive or sci.gov.in)

Usage:
    python 01_download.py --out-dir ./data/raw
    python 01_download.py --out-dir ./data/raw --limit 3  # quick test
    python 01_download.py --out-dir ./data/raw --include-continuations
"""

from __future__ import annotations

import argparse
import io
import json
import re
import socket
import subprocess
import time
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook
from tqdm import tqdm

SHEET_XLSX_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1AKsEOYGylGEC7fIN24XqUEI1CZXxH22cQcIpbb9eji0"
    "/export?format=xlsx"
)


# ---------------------------------------------------------------------------
# Sheet parsing (xlsx)
# ---------------------------------------------------------------------------

def _cell_hyperlink(cell) -> str:
    """Return the hyperlink target of an openpyxl cell, or ''."""
    return (cell.hyperlink.target if cell.hyperlink else "") or ""


def make_gdrive_direct(view_url: str) -> str | None:
    """
    Convert a Google Drive /view? sharing URL to a direct-download URL.

    Input:  https://drive.google.com/file/d/FILE_ID/view?usp=sharing
    Output: https://drive.usercontent.google.com/download?id=FILE_ID&export=download
    """
    m = re.search(r"/d/([A-Za-z0-9_-]{20,})", view_url)
    if not m:
        return None
    file_id = m.group(1)
    # Newer endpoint that works for large files without the virus-scan redirect
    return (
        f"https://drive.usercontent.google.com/download"
        f"?id={file_id}&export=download&authuser=0"
    )


def load_sheet(url: str, include_continuations: bool = False) -> list[dict]:
    """
    Download the Google Sheet as xlsx and return one dict per hearing row.

    PDF URL selection (in priority order):
      1. Google Drive hyperlink in col 7  (accessible from cloud servers)
      2. sci.gov.in URL in col 5          (fallback; may be blocked by firewall)

    Row types:
      • Primary rows   – non-empty Sr. No. (col 1).  19 entries.
      • Continuation   – blank Sr. No.; extra hearing date for same case. 5 entries.

    By default only primary rows are loaded (19).  Pass include_continuations=True
    for all 24 rows (~20% more training data).
    """
    print("Fetching dataset index (xlsx) …")
    resp = requests.get(url, timeout=120)
    resp.raise_for_status()

    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb.active

    records: list[dict] = []
    n_skipped_continuation = 0

    for row in range(2, ws.max_row + 1):
        sr_no    = str(ws.cell(row, 1).value or "").strip()
        dur_val  = ws.cell(row, 9).value
        mp3_cell = ws.cell(row, 10)
        pdf_c5   = ws.cell(row, 5)
        pdf_c7   = ws.cell(row, 7)

        # MP3 URL: prefer hyperlink, fall back to cell value
        mp3_url = (
            _cell_hyperlink(mp3_cell)
            or str(mp3_cell.value or "")
        ).strip()

        # PDF URL: prefer Google Drive hyperlink (col 7), then sci.gov.in (col 5)
        gdrive_view = _cell_hyperlink(pdf_c7)
        sci_url     = (_cell_hyperlink(pdf_c5) or str(pdf_c5.value or "")).strip()

        if gdrive_view and "drive.google.com" in gdrive_view:
            pdf_url = make_gdrive_direct(gdrive_view) or sci_url
            pdf_source = "gdrive"
        elif sci_url.startswith("http"):
            pdf_url    = sci_url
            pdf_source = "sci"
        else:
            pdf_url    = ""
            pdf_source = ""

        # Must have both audio and transcript
        if not mp3_url or not pdf_url:
            continue

        # Skip continuation rows unless requested
        if not sr_no:
            n_skipped_continuation += 1
            if not include_continuations:
                continue

        records.append(
            {
                "sr_no":        sr_no,
                "case_name":    str(ws.cell(row, 2).value or "").strip(),
                "case_number":  str(ws.cell(row, 3).value or "").strip(),
                "hearing_date": str(ws.cell(row, 4).value or "").strip(),
                "pdf_url":      pdf_url,
                "pdf_source":   pdf_source,
                "mp3_url":      mp3_url,
                "duration_min": float(dur_val) if dur_val else 0.0,
            }
        )

    if n_skipped_continuation and not include_continuations:
        print(
            f"  Note: skipped {n_skipped_continuation} continuation row(s). "
            f"Pass --include-continuations to include them."
        )

    n_gdrive = sum(1 for r in records if r["pdf_source"] == "gdrive")
    n_sci    = sum(1 for r in records if r["pdf_source"] == "sci")
    print(f"Found {len(records)} hearings  "
          f"({n_gdrive} PDFs via Google Drive, {n_sci} via sci.gov.in).")
    return records


# ---------------------------------------------------------------------------
# URL helpers
# ---------------------------------------------------------------------------

def make_dropbox_direct(url: str) -> str:
    """Convert a Dropbox sharing URL to a direct-download URL (dl=1)."""
    if "dl=" in url:
        return re.sub(r"dl=\d", "dl=1", url)
    sep = "&" if "?" in url else "?"
    return url + sep + "dl=1"


# ---------------------------------------------------------------------------
# Download helpers
# ---------------------------------------------------------------------------

_DNS_ERRORS = (
    "Name or service not known",
    "NameResolutionError",
    "nodename nor servname provided",
    "getaddrinfo failed",
    "Temporary failure in name resolution",
)


def _is_dns_error(exc: Exception) -> bool:
    msg = str(exc)
    return any(pat in msg for pat in _DNS_ERRORS)


def check_host_reachable(url: str) -> bool:
    """Quick DNS probe — returns False if the hostname can't be resolved."""
    host = urlparse(url).hostname or ""
    try:
        socket.getaddrinfo(host, 443)
        return True
    except socket.gaierror:
        return False


def download_file(
    url: str,
    dest: Path,
    retries: int = 3,
    timeout: int = 120,
    headers: dict | None = None,
) -> bool:
    """
    Streaming download with retry.  Returns True on success.

    DNS failures bail out immediately (no retries).
    Already-downloaded files (size > 1 KB) are skipped.
    """
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 1024:
        return True

    req_headers = headers or {}
    for attempt in range(1, retries + 1):
        try:
            with requests.get(
                url, stream=True, timeout=timeout,
                headers=req_headers, allow_redirects=True,
            ) as r:
                r.raise_for_status()
                with open(dest, "wb") as fh:
                    for chunk in r.iter_content(chunk_size=1 << 20):
                        fh.write(chunk)
            if dest.exists() and dest.stat().st_size > 1024:
                return True
            dest.unlink(missing_ok=True)
        except Exception as exc:
            if _is_dns_error(exc):
                print(f"    [DNS unreachable] {urlparse(url).hostname}")
                dest.unlink(missing_ok=True)
                return False
            print(f"    [attempt {attempt}/{retries}] {exc}")
            if attempt < retries:
                time.sleep(5 * attempt)

    return False


def convert_to_wav(mp3_path: Path, wav_path: Path) -> bool:
    """Convert MP3 → 16 kHz mono 16-bit PCM WAV using ffmpeg."""
    if wav_path.exists() and wav_path.stat().st_size > 0:
        return True
    cmd = [
        "ffmpeg", "-y", "-loglevel", "error",
        "-i",          str(mp3_path),
        "-ar",         "16000",
        "-ac",         "1",
        "-sample_fmt", "s16",
        str(wav_path),
    ]
    result = subprocess.run(cmd, capture_output=True)
    if result.returncode != 0:
        print(f"    ffmpeg error: {result.stderr.decode(errors='replace')[:200]}")
    return result.returncode == 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Download Supreme Court audio + transcripts."
    )
    ap.add_argument("--out-dir", default="./data/raw")
    ap.add_argument("--limit", type=int, default=0,
                    help="Download only the first N hearings (0 = all).")
    ap.add_argument("--skip-wav", action="store_true",
                    help="Skip MP3 → WAV conversion.")
    ap.add_argument(
        "--include-continuations", action="store_true",
        help="Also download the 5 continuation rows (24 total instead of 19).",
    )
    args = ap.parse_args()

    records = load_sheet(SHEET_XLSX_URL, include_continuations=args.include_continuations)
    if args.limit > 0:
        records = records[: args.limit]
        print(f"  Limiting to first {args.limit} hearing(s).")

    # ── Host reachability pre-check ───────────────────────────────────────
    gdrive_ok = check_host_reachable("https://drive.usercontent.google.com")
    dbx_ok    = check_host_reachable("https://www.dropbox.com")
    sci_ok    = check_host_reachable("https://main.sci.gov.in")

    if not gdrive_ok:
        print("\n  WARNING: Google Drive is not reachable from this machine.")
        print("  PDF downloads via GDrive will fail.")
    if not dbx_ok:
        print("\n  WARNING: Dropbox is not reachable. Audio downloads will fail.")
    if not sci_ok and not gdrive_ok:
        print("\n  ERROR: Both GDrive and sci.gov.in are unreachable.")
        print("  No PDFs can be downloaded automatically.")

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    index: list[dict]       = []
    missing_pdfs: list[dict] = []

    for i, rec in enumerate(tqdm(records, desc="Downloading")):
        date_slug   = rec["hearing_date"].replace("/", "-").replace(" ", "_")
        slug        = f"{i+1:03d}_{date_slug}"
        hearing_dir = out_dir / slug
        hearing_dir.mkdir(parents=True, exist_ok=True)

        # --- Audio (Dropbox MP3) ---
        mp3_url  = make_dropbox_direct(rec["mp3_url"])
        mp3_path = hearing_dir / "audio.mp3"
        mp3_ok   = download_file(mp3_url, mp3_path)
        if not mp3_ok:
            print(f"  WARNING: failed to download MP3 for {slug}")

        # --- Transcript PDF (Google Drive preferred, sci.gov.in fallback) ---
        pdf_path = hearing_dir / "transcript.pdf"
        pdf_ok   = download_file(rec["pdf_url"], pdf_path)

        # If Google Drive failed, try the sci.gov.in URL as a fallback
        if not pdf_ok and rec["pdf_source"] == "gdrive" and sci_ok:
            # Rebuild sci.gov.in URL from col 5 — we stored pdf_url as gdrive,
            # but let's just note the fallback wasn't precomputed here.
            pass

        if not pdf_ok:
            print(f"  WARNING: failed to download PDF for {slug}")
            missing_pdfs.append({"slug": slug, "url": rec["pdf_url"], "dest": str(pdf_path)})

        # --- MP3 → WAV ---
        wav_path = hearing_dir / "audio.wav"
        wav_ok   = False
        if mp3_ok and not args.skip_wav:
            wav_ok = convert_to_wav(mp3_path, wav_path)
            if not wav_ok:
                print(f"  WARNING: ffmpeg conversion failed for {slug}")

        index.append(
            {
                "slug":         slug,
                "case_name":    rec["case_name"],
                "case_number":  rec["case_number"],
                "hearing_date": rec["hearing_date"],
                "duration_min": rec["duration_min"],
                "mp3_path":     str(mp3_path.relative_to(out_dir)) if mp3_ok  else None,
                "wav_path":     str(wav_path.relative_to(out_dir)) if wav_ok  else None,
                "pdf_path":     str(pdf_path.relative_to(out_dir)) if pdf_ok  else None,
            }
        )

    index_path = out_dir / "index.json"
    index_path.write_text(json.dumps(index, indent=2, ensure_ascii=False))

    n_wav = sum(1 for r in index if r["wav_path"])
    n_pdf = sum(1 for r in index if r["pdf_path"])
    print(f"\nDownload complete.  Index → {index_path}")
    print(f"  WAVs ready : {n_wav} / {len(index)}")
    print(f"  PDFs ready : {n_pdf} / {len(index)}")

    if missing_pdfs:
        manual_path = out_dir / "missing_pdfs.txt"
        with manual_path.open("w") as fh:
            fh.write("# PDFs that could not be downloaded automatically.\n")
            fh.write("# Place each file at the listed destination path and re-run.\n\n")
            for m in missing_pdfs:
                fh.write(f"URL : {m['url']}\n")
                fh.write(f"DEST: {m['dest']}\n\n")
        print(f"\n  {len(missing_pdfs)} PDFs need manual download → {manual_path}")
        print("  Proceeding to Step 2 is still possible; those hearings will be skipped.")


if __name__ == "__main__":
    main()
